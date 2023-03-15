import json

import requests
import openpyxl

def get_product_info(product_name):
    # 定义请求头信息
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36',
    'Referer': 'https://www.jd.com/',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    }

    # 定义 cookie 值
    cookies = {
        'jd.utm': 'direct',
        'shshshfp': '',
        'shshshfpa': '',
        'shshshfpb': '',
        'shshshsID': '',
        '3AB9D23F7A4B3C9B': '',
        '__jdu': '',
        '__jdv': '',
        '__jda': '',
        '__jdb': '',
        '__jdc': '',
        'ipLoc-djd': '',
        'ipLoc-djd-p': '',
        'TrackID': '',
        'pinId': '',
        'shshshfpzc': '',
        'shshshsID': '',
        'shshshfpb': '',
        'shshshfp': '',
        'user-key': '',
    }

    # 创建 session，设置头信息和 cookie 值
    s = requests.Session()
    s.headers.update(headers)
    s.cookies.update(cookies)

    # 获取京东搜索结果页
    url1 = f'https://search.jd.com/Search?keyword={product_name}&enc=utf-8'
    response1 = s.get(url1)

    # 解析搜索结果页，获取商品链接
    item_url = ''
    print(response1.text)
    for line in response1.text.split('\n'):

        if 'J_goodsList' in line:
            json_str = line[line.index('{'):-2]
            resp_json = json.loads(json_str)
            item_list = resp_json['wareInfo']
            item_url = item_list[0]['productDetailUrl']
            break

    # 获取商品详情页
    response2 = s.get(item_url)

    # 解析商品详情页，获取商品名称和价格
    product_name = ''
    product_price = ''
    for line in response2.text.split('\n'):
        if 'p-name' in line:
            product_name = line.strip()[22:-7]
        if 'price J-p-' in line:
            begin_index = line.index('price J-p-') + 11
            end_index = line.index('"', begin_index)
            product_price = line[begin_index:end_index]
            break

    # 输出商品信息
    print(f'商品名称：{product_name}')
    print(f'商品价格：{product_price}')

    # 将商品信息写入excel表中
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '商品信息'
    ws['A1'] = '名称'
    ws['B1'] = '价格'
    ws['A2'] = product_name
    ws['B2'] = product_price
    wb.save(f'{product_name}.xlsx')

if __name__=='__main__':
    get_product_info('福通')